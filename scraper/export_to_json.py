import openpyxl
import json
import os

def read_excel(filepath, source_name):
    if not os.path.exists(filepath):
        return []
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    headers = [str(c.value) if c.value else "" for c in ws[1]]
    
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row): continue
        r = dict(zip(headers, ["" if v is None else v for v in row]))
        
        # Convert datetime objects to string if necessary
        deadline = r.get("Deadline", r.get("Posted", ""))
        if hasattr(deadline, "isoformat"):
            deadline = deadline.isoformat().split("T")[0]
            
        data.append({
            "id": str(r.get("Job ID", r.get("EJM ID", ""))),
            "source": str(r.get("Source", source_name)),
            "institution": str(r.get("Institution", r.get("University", ""))),
            "title": str(r.get("Position Title", "")),
            "country": str(r.get("Country", "")),
            "score": r.get("Score", 0),
            "deadline": str(deadline),
            "link": str(r.get("Apply Link", r.get("Application Link", ""))),
            "status": str(r.get("Status", "Not started"))
        })
    return data

def main():
    ejm = read_excel("uni_list.xlsx", "EJM")
    linkedin = read_excel("linkedin_list.xlsx", "LinkedIn")
    
    all_data = ejm + linkedin
    all_data.sort(key=lambda x: (-float(x["score"]) if str(x["score"]).replace('.','',1).isdigit() else 0))
    
    with open("../data.json", "w") as f:
        json.dump(all_data, f, indent=2)

if __name__ == "__main__":
    main()
