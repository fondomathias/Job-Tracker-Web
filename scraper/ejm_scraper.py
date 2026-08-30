#!/usr/bin/env python3
"""
Combined job-market scraper: econjobmarket.org (EJM) + European Job Market
for Economists (EJME, europeanjobmarketofeconomists.org).

Updates ../uni_list.xlsx with positions matching Paritosh's profile
(econometrics + close fields; Europe/India/Canada; Assistant Professor and
Postdoc; PLUS research institutions & central banks -- any source -- provided
non-Europeans are not explicitly excluded).

Run manually:
    python ejm_scraper.py            # full run (both sources)
    python ejm_scraper.py --no-profs # skip professor matching (faster)
    python ejm_scraper.py --no-ejme  # EJM feed only
    python ejm_scraper.py --dry-run  # show what would change, don't write

How the two sources combine
----------------------------
EJME listings are auto-pulled from econjobmarket.org, so every EJME ad has
an EJM position ID (linked on its detail page). The EJM public JSON feed is
the data backbone; EJME is crawled to (a) tag rows participating in EJME
coordination/signalling ("Source" column: EJM / EJME / EJM + EJME) and
(b) admit research-institution/central-bank ads that the university-oriented
type filter would drop.

Eligibility: every kept ad's text is screened for citizenship restrictions.
Explicit EU/EEA-only requirements -> excluded. Ambiguous phrases (work
permit, right to work, ...) -> kept, flagged in "Eligibility (non-EU)" with
the matched phrase quoted. No signal -> "No restriction found".

The Excel file is never overwritten destructively: rows keyed by EJM ID,
manual edits preserved, new rows highlighted, sorted by deadline.
"""

import argparse
import datetime as dt
import os
import re
import sys
import time
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import config

BASE = "https://econjobmarket.org"
FEED_URL = "https://backend.econjobmarket.org/data/zz_public/json/Ads"
EJME_BASE = "https://www.europeanjobmarketofeconomists.org"

HEADERS = [
    "EJM ID", "Source", "University", "Department", "Position Title", "Type",
    "Field(s)", "Priority", "Score", "Country", "City",
    "Eligibility (non-EU)", "Deadline", "Expected Answer / Interviews",
    "Start Date", "Duration", "Degree Required", "Documents Required",
    "# LoRs", "Application Link", "Dept Website",
    "Prof 1", "Prof 1 Link", "Prof 2", "Prof 2 Link", "Prof 3", "Prof 3 Link",
    "Date First Seen", "New?", "Status", "Notes",
]
SCRAPER_OWNED = {
    "Source", "University", "Department", "Position Title", "Type",
    "Field(s)", "Priority", "Score", "Country", "City",
    "Eligibility (non-EU)", "Deadline", "Expected Answer / Interviews",
    "Start Date", "Duration", "Degree Required", "Documents Required",
    "# LoRs", "Application Link", "Dept Website",
}
COL_WIDTHS = {
    "EJM ID": 8, "Source": 12, "University": 28, "Department": 22,
    "Position Title": 32, "Type": 18, "Field(s)": 28, "Priority": 9,
    "Score": 7, "Country": 14, "City": 13, "Eligibility (non-EU)": 26,
    "Deadline": 12, "Expected Answer / Interviews": 28, "Start Date": 12,
    "Duration": 15, "Degree Required": 13, "Documents Required": 32,
    "# LoRs": 7, "Application Link": 38, "Dept Website": 30,
    "Prof 1": 20, "Prof 1 Link": 30, "Prof 2": 20, "Prof 2 Link": 30,
    "Prof 3": 20, "Prof 3 Link": 30, "Date First Seen": 13, "New?": 6,
    "Status": 13, "Notes": 28,
}

DOC_PATTERNS = {
    "Cover letter": r"cover letter|letter of application|motivation letter",
    "CV": r"\bcv\b|curriculum vitae|r[e\xe9]sum[e\xe9]",
    "Job market paper": r"job market paper|jmp\b",
    "Research statement": r"research statement|statement of research|research proposal",
    "Teaching statement": r"teaching statement|statement of teaching|teaching philosophy",
    "Writing/research sample": r"writing sample|research sample|sample of (recent )?research|research papers?",
    "Transcripts": r"transcript|diplomas?\b",
    "Diversity statement": r"diversity statement|statement of diversity",
    "Publication list": r"list of publications|publication list",
}

EJME_TYPE_NAMES = [
    "Assistant Professor", "Associate Professor", "Full Professor",
    "Tenured Professor", "Untenured Professor", "Visiting Assistant Professor",
    "Visiting Associate Professor", "Visiting Professor",
    "Postdoctoral Scholar", "Lecturer", "Other academic", "Other nonacademic",
    "Consultant", "Doctoral student", "Research Assistant",
]


# --------------------------------------------------------------------------
# small utilities
# --------------------------------------------------------------------------
def log(msg):
    print(f"[{dt.datetime.now():%H:%M:%S}] {msg}", flush=True)


def polite_get(session, url, **kw):
    time.sleep(config.REQUEST_DELAY_SECONDS)
    return session.get(url, timeout=config.TIMEOUT, **kw)


def load_credentials():
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        config.CREDENTIALS_FILE)
    creds = {}
    if os.path.exists(path):
        with open(path, encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, _, v = line.partition("=")
                    creds[k.strip()] = v.strip()
    return creds.get("EJM_EMAIL"), creds.get("EJM_PASSWORD")


def parse_date(text):
    """datetime.date from '30 Nov 2026', '2026-06-30', '05/31/2026', ..."""
    if not text:
        return None
    if isinstance(text, dt.datetime):
        return text.date()
    if isinstance(text, dt.date):
        return text
    text = str(text).strip()
    for fmt in ("%Y-%m-%d", "%d %b %Y", "%d %B %Y", "%b %d, %Y",
                "%B %d, %Y", "%m/%d/%Y"):
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    m = re.search(r"(\d{1,2})\s+([A-Za-z]{3,9})\s+(\d{4})", text)
    if m:
        for fmt in ("%d %b %Y", "%d %B %Y"):
            try:
                return dt.datetime.strptime(" ".join(m.groups()), fmt).date()
            except ValueError:
                continue
    m = re.search(r"(\d{2}/\d{2}/\d{4})", text)
    if m:
        try:
            return dt.datetime.strptime(m.group(1), "%m/%d/%Y").date()
        except ValueError:
            pass
    return None


def is_research_inst(name):
    n = (name or "").lower()
    return any(kw in n for kw in config.RESEARCH_INST_KEYWORDS)


# --------------------------------------------------------------------------
# login (econjobmarket.org; optional)
# --------------------------------------------------------------------------
def make_session():
    s = requests.Session()
    s.headers.update({"User-Agent": config.USER_AGENT})
    return s


def try_login(session):
    email, password = load_credentials()
    if not email or not password:
        log("No credentials.env found (or empty) -- continuing without login.")
        return False
    try:
        r = polite_get(session, f"{BASE}/login")
        soup = BeautifulSoup(r.text, "lxml")
        token_input = soup.find("input", {"name": "_token"})
        data = {"email": email, "password": password}
        if token_input:
            data["_token"] = token_input.get("value", "")
        time.sleep(config.REQUEST_DELAY_SECONDS)
        r = session.post(f"{BASE}/login", data=data, timeout=config.TIMEOUT,
                         allow_redirects=True)
        logged_in = "logout" in r.text.lower()
        log("Login successful." if logged_in
            else "Login may have failed -- continuing anyway.")
        return logged_in
    except requests.RequestException as exc:
        log(f"Login skipped (network issue: {exc}).")
        return False


# --------------------------------------------------------------------------
# eligibility screening (non-European applicants)
# --------------------------------------------------------------------------
def eligibility_check(text):
    """Return (status_string, keep_bool). Hard EU/EEA-only rules -> exclude;
    ambiguous phrases -> keep but flag with the quoted match."""
    tl = " ".join((text or "").lower().split())
    for pat in config.HARD_RESTRICT_PATTERNS:
        m = re.search(pat, tl)
        if m:
            return f"RESTRICTED: \"{m.group(0)[:60]}\"", False
    for pat in config.SOFT_RESTRICT_PATTERNS:
        m = re.search(pat, tl)
        if m:
            return f"CHECK: \"{m.group(0)[:60]}\"", True
    return "No restriction found", True


# --------------------------------------------------------------------------
# EJM feed: fetch + pre-filter
# --------------------------------------------------------------------------
def fetch_feed(session):
    log("Fetching EJM public ads feed ...")
    r = polite_get(session, FEED_URL)
    r.raise_for_status()
    ads = r.json()
    log(f"Feed contains {len(ads)} active ads.")
    return ads


def in_bounding_box(lat, lon):
    if lat is None or lon is None:
        return True
    for (a, b, c, d) in config.BOUNDING_BOXES:
        if a <= lat <= b and c <= lon <= d:
            return True
    return False


def classify_type(position_types, university=""):
    """Map EJM position types; universities keep only KEEP_TYPES, research
    institutions / central banks also admit RESEARCH_EXTRA_TYPES."""
    names = [p.get("name", "").lower() for p in (position_types or [])]
    mapped = [config.TYPE_MAP.get(n) for n in names if config.TYPE_MAP.get(n)]
    for want in config.KEEP_TYPES:
        if want in mapped:
            return want, " | ".join(sorted(set(mapped))), True
    if is_research_inst(university):
        for want in config.RESEARCH_EXTRA_TYPES:
            if want in mapped:
                return (f"Researcher ({want})",
                        " | ".join(sorted(set(mapped))), True)
    return (mapped[0] if mapped else "Other",
            " | ".join(sorted(set(mapped))), False)


def keyword_score(text):
    tl = (text or "").lower()
    return sum(w for kw, w in config.KEYWORD_WEIGHTS.items() if kw in tl)


def score_ad(ad):
    score = 0
    cats = [c.get("name", "").lower() for c in (ad.get("categories") or [])]
    for c in cats:
        if any(p in c for p in config.PRIMARY_CATEGORIES):
            score += 10
        elif any(sc in c for sc in config.SECONDARY_CATEGORIES):
            score += 4
    score += keyword_score(ad.get("adtitle", "") + " " + ad.get("adtext", ""))
    return score


def prefilter(ads):
    keep = []
    for ad in ads:
        ejm_id = (ad.get("url") or "").rstrip("/").rsplit("/", 1)[-1]
        if not ejm_id.isdigit():
            continue
        if not in_bounding_box(ad.get("latitude"), ad.get("longitude")):
            continue
        score = score_ad(ad)
        if score < config.MIN_SCORE:
            continue
        university = (ad.get("name") or "").strip()
        ad_type, all_types, type_ok = classify_type(
            ad.get("position_types"), university)
        if not type_ok:
            continue
        keep.append({
            "id": ejm_id,
            "score": score,
            "type": ad_type,
            "all_types": all_types,
            "title": ad.get("adtitle", "").strip(),
            "university": university,
            "department": (ad.get("department") or "").strip(),
            "fields": " | ".join(c.get("name", "")
                                 for c in (ad.get("categories") or [])),
            "deadline": parse_date(ad.get("deadline_date")),
            "url": ad.get("url", f"{BASE}/positions/{ejm_id}"),
            "adtext": ad.get("adtext", ""),
        })
    log(f"EJM feed: {len(keep)} ads survive geography-box + field + type "
        f"pre-filter.")
    return keep


# --------------------------------------------------------------------------
# EJME crawl: listing pages -> detail pages -> EJM position IDs
# --------------------------------------------------------------------------
def _ejme_parse_listing_page(html, page_url):
    """Return [{url,title,advertiser,location,types}] from one listing page."""
    soup = BeautifulSoup(html, "lxml")
    items, seen = [], set()
    for a in soup.find_all("a", href=True):
        if "/job-listing/" not in a["href"]:
            continue
        title = a.get_text(" ", strip=True)
        if not title or title.lower().startswith("read more"):
            continue
        href = urljoin(page_url, a["href"])
        if href in seen:
            continue
        seen.add(href)
        block, txt = a, ""
        for _ in range(6):
            block = block.parent
            if block is None:
                break
            txt = block.get_text("\n", strip=True)
            if "Advertiser" in txt:
                break
        adv = re.search(r"Advertiser\(s\):\s*\n?(.+)", txt)
        loc = re.search(r"Location:\s*\n?(.+)", txt)
        # keep only the local text near this listing (block may be too big)
        types = [t for t in EJME_TYPE_NAMES
                 if re.search(rf"(?<![\w]){re.escape(t)}(?![\w])", txt)]
        items.append({
            "url": href,
            "title": title,
            "advertiser": adv.group(1).strip() if adv else "",
            "location": loc.group(1).strip() if loc else "",
            "types": types,
        })
    return items


def ejme_collect(session):
    """Crawl EJME; return {ejm_id: listing_meta} for listings passing the
    cheap listing-level filters (country + type/research-institution)."""
    if not getattr(config, "EJME_ENABLED", True):
        return {}
    listings = []
    for p in range(config.EJME_MAX_PAGES):
        url = f"{config.EJME_LISTING_URL}?page={p}"
        try:
            r = polite_get(session, url)
            r.raise_for_status()
        except requests.RequestException as exc:
            log(f"EJME: stopping pagination at page {p} ({exc})")
            break
        items = _ejme_parse_listing_page(r.text, url)
        if not items:
            break
        listings.extend(items)
        if f"page={p + 1}" not in r.text:
            break
    log(f"EJME: {len(listings)} listings collected.")

    survivors = []
    for it in listings:
        country = it["location"].rsplit(",", 1)[-1].strip() \
            if it["location"] else ""
        it["country"] = country
        mapped = [config.TYPE_MAP.get(t.lower()) for t in it["types"]]
        mapped = [m for m in mapped if m]
        type_ok = (
            any(m in config.KEEP_TYPES for m in mapped)
            or (is_research_inst(it["advertiser"])
                and any(m in config.RESEARCH_EXTRA_TYPES for m in mapped))
            or not mapped          # unparsed -> decide at detail stage
        )
        geo_ok = (not country
                  or country in ("Multi-country", "Worldwide")
                  or country_ok(country))
        if type_ok and geo_ok:
            survivors.append(it)
    log(f"EJME: {len(survivors)} pass listing-level filters; "
        f"fetching detail pages ...")

    mapping = {}
    for it in survivors:
        try:
            r = polite_get(session, it["url"])
            r.raise_for_status()
        except requests.RequestException:
            continue
        m = re.search(r"econjobmarket\.org/positions/(\d+)", r.text)
        if m:
            mapping[m.group(1)] = it
    log(f"EJME: {len(mapping)} listings mapped to EJM position IDs.")
    return mapping


# --------------------------------------------------------------------------
# ad-page enrichment + final country filter
# --------------------------------------------------------------------------
AD_FIELD_PATTERNS = {
    "location": r"Location of job:\s*(.+)",
    "degree": r"Degree required:\s*(.+)",
    "start_date": r"Job start date:\s*(.+)",
    "duration": r"Job duration:\s*(.+)",
    "letters": r"Letters of reference required:\s*(\d+)",
    "deadline_page": r"Application deadline:\s*(.+)",
    "interviews": r"Interviews:\s*(.+)",
}


def parse_ad_page(session, ad):
    try:
        r = polite_get(session, ad["url"])
        r.raise_for_status()
    except requests.RequestException as exc:
        log(f"  ! could not fetch ad {ad['id']}: {exc}")
        ad.setdefault("eligibility", "")
        ad.setdefault("eligible_keep", True)
        return ad
    soup = BeautifulSoup(r.text, "lxml")
    text = soup.get_text("\n", strip=True)

    for key, pat in AD_FIELD_PATTERNS.items():
        m = re.search(pat, text)
        if m:
            ad[key] = m.group(1).strip()

    loc = ad.get("location", "")
    parts = [p.strip() for p in loc.split(",") if p.strip()]
    country = parts[-1] if parts else ""
    ad["country"] = country
    city = ""
    for p in reversed(parts[:-1]):
        if p.lower() == country.lower():
            continue
        if re.search(r"\d", p) and len(re.sub(r"[\d\s\-]", "", p)) <= 3:
            continue
        city = p
        break
    ad["city"] = city

    sites = []
    for a in soup.find_all("a", href=True):
        img = a.find("img")
        if img and "www" in (img.get("src") or ""):
            href = a["href"]
            if href.startswith("http") and "econjobmarket" not in href:
                sites.append(href)
    ad["dept_site"] = sites[0] if sites else ""

    body = (ad.get("adtext", "") + " " + text).lower()
    docs = [name for name, pat in DOC_PATTERNS.items()
            if re.search(pat, body)]
    ad["documents"] = ", ".join(docs)

    # eligibility screening on the full ad text
    ad["eligibility"], ad["eligible_keep"] = eligibility_check(body)

    # EJME-only ads arrive without feed metadata: score from page text,
    # including the category-style bonus feed ads get (the ad page lists
    # the position's fields, so scan the text for category names).
    if ad.get("ejme_only"):
        s = keyword_score(text)
        tl = text.lower()
        if any(p in tl for p in config.PRIMARY_CATEGORIES):
            s += 10
        elif any(sc in tl for sc in config.SECONDARY_CATEGORIES):
            s += 4
        ad["score"] = max(ad.get("score", 0), s)

    if not ad.get("deadline") and ad.get("deadline_page"):
        ad["deadline"] = parse_date(ad["deadline_page"])
    return ad


def country_ok(country):
    c = (country or "").lower().strip()
    return any(c == w.lower() for w in config.COUNTRY_WHITELIST)


# --------------------------------------------------------------------------
# professor matching (best-effort, shallow crawl)
# --------------------------------------------------------------------------
FACULTY_LINK_HINTS = ["people", "faculty", "staff", "members", "directory",
                      "professor", "researchers", "team", "personnel"]
NAME_RE = re.compile(r"^[A-Z\xc0-\xde][\w'\-.]+(\s+[A-Z\xc0-\xde][\w'\-.]+){1,3}$")


def same_domain(url, base_url):
    try:
        return urlparse(url).netloc in ("", urlparse(base_url).netloc)
    except ValueError:
        return False


def find_faculty_pages(session, dept_url):
    try:
        r = polite_get(session, dept_url)
        r.raise_for_status()
    except requests.RequestException:
        return []
    soup = BeautifulSoup(r.text, "lxml")
    scored = []
    for a in soup.find_all("a", href=True):
        label = (a.get_text(" ", strip=True) + " " + a["href"]).lower()
        hits = sum(1 for h in FACULTY_LINK_HINTS if h in label)
        if hits:
            url = urljoin(dept_url, a["href"])
            if same_domain(url, dept_url):
                scored.append((hits, url))
    scored.sort(key=lambda t: -t[0])
    seen, out = set(), []
    for _, url in scored:
        if url not in seen:
            seen.add(url)
            out.append(url)
        if len(out) >= config.PROF_MAX_FACULTY_PAGES:
            break
    return out


def harvest_profiles(session, faculty_url):
    try:
        r = polite_get(session, faculty_url)
        r.raise_for_status()
    except requests.RequestException:
        return []
    soup = BeautifulSoup(r.text, "lxml")
    people = []
    for a in soup.find_all("a", href=True):
        name = a.get_text(" ", strip=True)
        if 5 <= len(name) <= 60 and NAME_RE.match(name):
            url = urljoin(faculty_url, a["href"])
            if same_domain(url, faculty_url):
                people.append((name, url))
    seen, out = set(), []
    for name, url in people:
        if url not in seen:
            seen.add(url)
            out.append((name, url))
    return out


def score_profile(session, url):
    try:
        r = polite_get(session, url)
        r.raise_for_status()
    except requests.RequestException:
        return 0
    text = BeautifulSoup(r.text, "lxml").get_text(" ", strip=True).lower()
    return sum(text.count(kw) for kw in config.PROF_KEYWORDS)


def match_professors(session, dept_url):
    if not dept_url:
        return []
    log(f"  prof-matching on {dept_url} ...")
    candidates = []
    for fac_url in find_faculty_pages(session, dept_url):
        candidates.extend(harvest_profiles(session, fac_url))
        if len(candidates) >= config.PROF_MAX_PROFILES:
            break
    candidates = candidates[:config.PROF_MAX_PROFILES]
    ranked = []
    for name, url in candidates:
        s = score_profile(session, url)
        if s > 0:
            ranked.append((s, name, url))
    ranked.sort(key=lambda t: -t[0])
    return [(name, url) for _, name, url in ranked[:3]]


# --------------------------------------------------------------------------
# Excel update
# --------------------------------------------------------------------------
NEW_FILL = PatternFill("solid", start_color="FFF3C7")
HEAD_FILL = PatternFill("solid", start_color="1F4E79")
HEAD_FONT = Font(color="FFFFFF", bold=True)


def read_existing(path):
    rows = {}
    if not os.path.exists(path):
        return rows
    try:
        wb = load_workbook(path)
    except Exception:
        return rows
    ws = wb.active
    if ws.max_row < 2:
        return rows
    header = [c.value for c in ws[1]]
    if "EJM ID" not in header:
        return rows
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(header, row))
        key = str(rec.get("EJM ID") or "").strip()
        if key:
            rows[key] = rec
    return rows


def ad_to_record(ad, today):
    profs = list(ad.get("profs", []))
    profs += [("", "")] * (3 - len(profs))
    return {
        "EJM ID": ad["id"],
        "Source": ad.get("source", "EJM"),
        "University": ad["university"],
        "Department": ad["department"],
        "Position Title": ad["title"],
        "Type": ad["type"],
        "Field(s)": ad["fields"],
        "Priority": ("High" if ad["score"] >= config.PRIORITY_HIGH else
                     "Medium" if ad["score"] >= config.PRIORITY_MEDIUM
                     else "Low"),
        "Score": ad["score"],
        "Country": ad.get("country", ""),
        "City": ad.get("city", ""),
        "Eligibility (non-EU)": ad.get("eligibility", ""),
        "Deadline": ad.get("deadline"),
        "Expected Answer / Interviews": ad.get("interviews", ""),
        "Start Date": ad.get("start_date", ""),
        "Duration": ad.get("duration", ""),
        "Degree Required": ad.get("degree", ""),
        "Documents Required": ad.get("documents", ""),
        "# LoRs": ad.get("letters", ""),
        "Application Link": ad["url"],
        "Dept Website": ad.get("dept_site", ""),
        "Prof 1": profs[0][0], "Prof 1 Link": profs[0][1],
        "Prof 2": profs[1][0], "Prof 2 Link": profs[1][1],
        "Prof 3": profs[2][0], "Prof 3 Link": profs[2][1],
        "Date First Seen": today,
        "New?": "NEW",
        "Status": "Not started",
        "Notes": "",
    }


def merge(existing, fresh):
    merged = dict(existing)
    for col in HEADERS:
        new_val = fresh.get(col)
        if col in SCRAPER_OWNED and new_val not in (None, ""):
            merged[col] = new_val
    merged["New?"] = ""
    for col in HEADERS:
        merged.setdefault(col, "")
    return merged


def write_workbook(path, records, new_ids):
    records.sort(key=lambda r: (r.get("Deadline") is None,
                                r.get("Deadline") or dt.date.max,
                                r.get("University") or ""))
    wb = Workbook()
    ws = wb.active
    ws.title = "Positions"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
    for rec in records:
        row = []
        for h in HEADERS:
            v = rec.get(h, "")
            if isinstance(v, dt.date):
                v = v.isoformat()
            row.append(v)
        ws.append(row)
        if str(rec.get("EJM ID")) in new_ids:
            for cell in ws[ws.max_row]:
                cell.fill = NEW_FILL
    for i, h in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(i)].width = \
            COL_WIDTHS.get(h, 16)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"
    dv = DataValidation(
        type="list",
        formula1='"Not started,Drafting,Applied,Interview,Flyout,Offer,'
                 'Rejected,Withdrawn"',
        allow_blank=True)
    ws.add_data_validation(dv)
    status_col = get_column_letter(HEADERS.index("Status") + 1)
    dv.add(f"{status_col}2:{status_col}{max(ws.max_row, 2)}")
    wb.save(path)


# --------------------------------------------------------------------------
# main
# --------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description="EJM + EJME position scraper")
    ap.add_argument("--no-profs", action="store_true",
                    help="skip professor matching (much faster)")
    ap.add_argument("--no-ejme", action="store_true",
                    help="skip the EJME source")
    ap.add_argument("--dry-run", action="store_true",
                    help="don't write the Excel file")
    args = ap.parse_args()

    here = os.path.dirname(os.path.abspath(__file__))
    xlsx_path = os.path.normpath(os.path.join(here, config.XLSX_PATH))
    today = dt.date.today().isoformat()

    session = make_session()
    try_login(session)

    # ---- source 2 first (cheap): EJME tag map ----
    ejme = {}
    if not args.no_ejme:
        try:
            ejme = ejme_collect(session)
        except Exception as exc:
            log(f"EJME crawl failed ({exc}) -- continuing with EJM only.")

    # ---- source 1: EJM feed ----
    ads = prefilter(fetch_feed(session))
    feed_ids = {ad["id"] for ad in ads}

    # EJME listings absent from the filtered feed set: admit them as
    # EJME-sourced candidates (typically research institutions / central
    # banks, or ads the field-score filter missed but EJME deems relevant).
    for eid, meta in ejme.items():
        if eid in feed_ids:
            continue
        ptypes = [{"name": t} for t in meta.get("types", [])]
        ad_type, all_types, type_ok = classify_type(ptypes,
                                                    meta.get("advertiser", ""))
        if ptypes and not type_ok:
            continue
        ads.append({
            "id": eid,
            "score": 0,                      # scored from ad page text later
            "type": ad_type if ptypes else "Other",
            "all_types": all_types,
            "title": meta.get("title", ""),
            "university": meta.get("advertiser", ""),
            "department": "",
            "fields": "",
            "deadline": None,
            "url": f"{BASE}/positions/{eid}",
            "adtext": "",
            "ejme_only": True,
        })

    existing = read_existing(xlsx_path)
    log(f"Tracker currently holds {len(existing)} positions.")

    enriched, new_ids = [], set()
    for i, ad in enumerate(ads, 1):
        log(f"[{i}/{len(ads)}] {ad['university']} -- {ad['title'][:60]}")
        ad = parse_ad_page(session, ad)
        if not country_ok(ad.get("country")):
            log(f"  skipped (country: {ad.get('country') or 'unknown'})")
            continue
        if ad.get("ejme_only") and ad["score"] < config.EJME_MIN_SCORE:
            log(f"  skipped (EJME ad, field score {ad['score']} < "
                f"{config.EJME_MIN_SCORE})")
            continue
        if not ad.get("eligible_keep", True):
            log(f"  EXCLUDED ({ad.get('eligibility')})")
            continue
        if (not config.INCLUDE_PAST_DEADLINES and ad.get("deadline")
                and ad["deadline"] < dt.date.today()):
            log(f"  skipped (deadline already passed: {ad['deadline']})")
            continue
        ad["source"] = ("EJME" if ad.get("ejme_only")
                        else "EJM + EJME" if ad["id"] in ejme else "EJM")
        is_new = ad["id"] not in existing
        if is_new:
            new_ids.add(ad["id"])
        run_profs = (config.PROF_ENABLED and not args.no_profs and
                     (is_new or not existing.get(ad["id"], {}).get("Prof 1")))
        if run_profs:
            try:
                ad["profs"] = match_professors(session, ad.get("dept_site"))
            except Exception as exc:
                log(f"  prof matching failed: {exc}")
                ad["profs"] = []
        enriched.append(ad)

    final, seen = [], set()
    for ad in enriched:
        rec = ad_to_record(ad, today)
        if ad["id"] in existing:
            rec = merge(existing[ad["id"]], rec)
        else:
            rec["New?"] = "NEW"
        final.append(rec)
        seen.add(ad["id"])
    for ejm_id, rec in existing.items():
        if ejm_id not in seen:
            rec = dict(rec)
            if not rec.get("Notes"):
                rec["Notes"] = "No longer listed / filtered out"
            for col in HEADERS:
                rec.setdefault(col, "")
            final.append(rec)

    for rec in final:
        if not isinstance(rec.get("Deadline"), dt.date):
            rec["Deadline"] = parse_date(rec.get("Deadline"))
        elif isinstance(rec.get("Deadline"), dt.datetime):
            rec["Deadline"] = rec["Deadline"].date()

    n_tagged = sum(1 for r in final if r.get("Source") == "EJM + EJME")
    log(f"Result: {len(final)} rows total, {len(new_ids)} new, "
        f"{n_tagged} tagged EJM + EJME.")
    if args.dry_run:
        for rec in final:
            if str(rec.get("EJM ID")) in new_ids:
                log(f"  NEW [{rec.get('Source')}]: {rec['University']} -- "
                    f"{rec['Position Title']} (deadline {rec.get('Deadline')})"
                    f" | {rec.get('Eligibility (non-EU)', '')}")
        log("Dry run - Excel file not modified.")
        return

    write_workbook(xlsx_path, final, new_ids)
    log(f"Updated {xlsx_path}")
    log("New rows highlighted; sheet sorted by deadline. Check any "
        "'CHECK:' entries in the Eligibility column manually.")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        sys.exit("\nInterrupted.")
