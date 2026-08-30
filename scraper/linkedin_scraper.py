#!/usr/bin/env python3
"""
LinkedIn job scraper -- guest (logged-out) endpoints only.

Searches LinkedIn's public job search for economics/econometrics academic
positions (Europe / India / Canada) and writes results to
linkedin_list.xlsx IN THIS FOLDER (kept separate from ../uni_list.xlsx so
the EJM scraper never touches these rows and vice versa).

    python linkedin_scraper.py               # postings from the last 7 days
    python linkedin_scraper.py --days 30     # wider window (first run)
    python linkedin_scraper.py --dry-run     # preview, don't write Excel

Why guest endpoints:
- No login, no API key, no paid tools. Your LinkedIn account is never
  involved, so it cannot be flagged or restricted.
- They see ALL public job posts -- coverage is independent of your
  connections.
- LinkedIn throttles by IP (HTTP 429): the scraper waits and retries, and
  saves whatever it collected if throttling persists. Running once a day
  at the default politeness settings is normally fine.

Honest limitations:
- LinkedIn changes its markup without notice; the parser is written
  tolerantly but may need patching occasionally.
- Job posts carry no application deadline; the sheet is sorted by
  relevance score, with the posting date as the urgency signal.
- Some academic jobs are posted only via university portals, never on
  LinkedIn. This complements, not replaces, the EJM/EJME scraper.
"""

import argparse
import datetime as dt
import os
import random
import re
import sys
import time
from urllib.parse import quote, urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import config  # read-only: reuses COUNTRY_WHITELIST, KEYWORD_WEIGHTS,
               # eligibility patterns, USER_AGENT from the EJM config

# ======================== LinkedIn-specific settings ========================
SEARCH_TERMS = [
    "assistant professor economics",
    "postdoctoral researcher economics",
    "postdoc economics",
    "postdoc econometrics",
    "research economist",
    "lecturer economics",
]
LOCATIONS = [
    "European Union",
    "Finland",
    "Sweden",
    "Denmark",
    "Estonia",
    "Latvia",
    "Lithuania",
    "Germany",
    "Netherlands",
    "Belgium",
    "France",
    "Spain",
    "Italy",
    "Portugal",
    "Norway",
    "United Kingdom",
    "Switzerland",
    "India",
    "Canada",
]
MAX_PAGES_PER_QUERY = 5        # 25 jobs/page 
MAX_DETAIL_FETCHES = 150       # cap on description downloads per run
DELAY_SECONDS = 3.0            # base politeness delay (jitter added)
THROTTLE_WAIT = 60             # wait after an HTTP 429 before retrying
MIN_SCORE = 8                  # LinkedIn is noisy; higher bar than EJM
PRIORITY_HIGH = 16
PRIORITY_MEDIUM = 11

# A job title must contain one of these (case-insensitive) before we spend
# a request on its description:
TITLE_HINTS = [
    "professor", "postdoc", "post-doc", "postdoctoral", "lecturer",
    "researcher", "research fellow", "fellow", "faculty", "economist",
    "econometrician", "scientist",
]
# Title/company bonuses on top of config.KEYWORD_WEIGHTS text scoring:
TITLE_BONUSES = {
    r"professor|postdoc|post-doc|lecturer|research fellow": 5,
    r"econometric": 6,
    r"economics|economist": 3,
}
COMPANY_BONUSES = {
    r"university|universit|school of economics|college": 4,
    r"bank|institute|institut|research": 3,
}

OUT_XLSX = "linkedin_list.xlsx"          # written next to this script
SEARCH_URL = ("https://www.linkedin.com/jobs-guest/jobs/api/"
              "seeMoreJobPostings/search")
DETAIL_URL = ("https://www.linkedin.com/jobs-guest/jobs/api/jobPosting/{}")

HEADERS_XLSX = [
    "Job ID", "Source", "Institution", "Position Title", "Location",
    "Country", "Posted", "Score", "Priority", "Eligibility (non-EU)",
    "Apply Link", "In uni_list?", "Date First Seen", "New?", "Status",
    "Notes",
]
SCRAPER_OWNED = {"Source", "Institution", "Position Title", "Location",
                 "Country", "Posted", "Score", "Priority",
                 "Eligibility (non-EU)", "Apply Link", "In uni_list?"}
COL_WIDTHS = {"Job ID": 12, "Source": 10, "Institution": 30,
              "Position Title": 38, "Location": 24, "Country": 14,
              "Posted": 11, "Score": 7, "Priority": 9,
              "Eligibility (non-EU)": 26, "Apply Link": 40,
              "In uni_list?": 12, "Date First Seen": 13, "New?": 6,
              "Status": 13, "Notes": 28}
# ===========================================================================


def log(msg):
    print(f"[{dt.datetime.now():%H:%M:%S}] {msg}", flush=True)


def polite_get(session, url, params=None):
    """GET with jittered delay; waits out one 429/999 throttle, then gives
    up on that URL (returns None) so the run can continue."""
    for attempt in (1, 2):
        time.sleep(DELAY_SECONDS + random.uniform(0, 1.5))
        try:
            r = session.get(url, params=params, timeout=25)
        except requests.RequestException as exc:
            log(f"  network error: {exc}")
            return None
        if r.status_code == 200:
            return r
        if r.status_code in (429, 999) and attempt == 1:
            log(f"  throttled (HTTP {r.status_code}); waiting "
                f"{THROTTLE_WAIT}s ...")
            time.sleep(THROTTLE_WAIT)
            continue
        log(f"  HTTP {r.status_code} for {url[:80]}")
        return None
    return None


def make_session():
    s = requests.Session()
    s.headers.update({
        "User-Agent": config.USER_AGENT.replace("EJM-personal-tracker",
                                                "").strip(),
        "Accept-Language": "en-US,en;q=0.9",
    })
    return s


# --------------------------------------------------------------- parsing --
def parse_cards(html, base="https://www.linkedin.com"):
    """Parse a seeMoreJobPostings fragment into job dicts."""
    soup = BeautifulSoup(html, "lxml")
    jobs = []
    for card in soup.select("[data-entity-urn*='jobPosting'], li"):
        urn = card.get("data-entity-urn", "")
        if "jobPosting" not in urn:
            inner = card.select_one("[data-entity-urn*='jobPosting']")
            if inner is None:
                continue
            card, urn = inner, inner.get("data-entity-urn", "")
        m = re.search(r"jobPosting:(\d+)", urn)
        if not m:
            continue
        jid = m.group(1)

        def txt(sel):
            el = card.select_one(sel)
            return el.get_text(" ", strip=True) if el else ""

        title = (txt("h3.base-search-card__title") or txt("h3") or
                 txt(".base-search-card__title"))
        company = (txt("h4.base-search-card__subtitle") or txt("h4") or
                   txt(".base-search-card__subtitle"))
        location = (txt(".job-search-card__location") or
                    txt(".base-search-card__metadata"))
        t = card.select_one("time[datetime]")
        posted = t.get("datetime", "") if t else ""
        a = card.select_one("a.base-card__full-link, a[href*='/jobs/view/']")
        link = urljoin(base, a["href"].split("?")[0]) if a and \
            a.get("href") else f"{base}/jobs/view/{jid}"
        jobs.append({"id": jid, "title": title, "company": company,
                     "location": location, "posted": posted, "link": link})
    # de-duplicate within the fragment
    seen, out = set(), []
    for j in jobs:
        if j["id"] not in seen:
            seen.add(j["id"])
            out.append(j)
    return out


def fetch_description(session, jid):
    r = polite_get(session, DETAIL_URL.format(jid))
    if r is None:
        return ""
    soup = BeautifulSoup(r.text, "lxml")
    el = (soup.select_one(".show-more-less-html__markup") or
          soup.select_one(".description__text") or soup)
    return el.get_text(" ", strip=True)


# --------------------------------------------------------------- scoring --
def title_ok(title):
    tl = (title or "").lower()
    return any(h in tl for h in TITLE_HINTS)


def country_of(location):
    parts = [p.strip() for p in (location or "").split(",") if p.strip()]
    return parts[-1] if parts else ""


def country_ok(country):
    c = (country or "").lower().strip()
    return any(c == w.lower() for w in config.COUNTRY_WHITELIST)


def score_job(job, description):
    score = 0
    tl = (job["title"] or "").lower()
    cl = (job["company"] or "").lower()
    for pat, w in TITLE_BONUSES.items():
        if re.search(pat, tl):
            score += w
    for pat, w in COMPANY_BONUSES.items():
        if re.search(pat, cl):
            score += w
    text = (tl + " " + (description or "").lower())
    for kw, w in config.KEYWORD_WEIGHTS.items():
        if kw in text:
            score += w
    return score


def eligibility_check(text):
    tl = " ".join((text or "").lower().split())
    for pat in config.HARD_RESTRICT_PATTERNS:
        m = re.search(pat, tl)
        if m:
            return f"RESTRICTED: \"{m.group(0)[:60]}\"", False
    for pat in config.SOFT_RESTRICT_PATTERNS:
        m = re.search(pat, tl)
        if m:
            return f"CHECK: \"{m.group(0)[:60]}\"", True
    return "No restriction found", True


# ----------------------------------------------------------------- excel --
NEW_FILL = PatternFill("solid", start_color="FFF3C7")
HEAD_FILL = PatternFill("solid", start_color="1F4E79")
HEAD_FONT = Font(color="FFFFFF", bold=True)


def read_existing(path):
    rows = {}
    if not os.path.exists(path):
        return rows
    try:
        wb = load_workbook(path)
    except Exception:
        return rows
    ws = wb.active
    header = [c.value for c in ws[1]]
    if "Job ID" not in header:
        return rows
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(header, row))
        key = str(rec.get("Job ID") or "").strip()
        if key:
            rows[key] = rec
    return rows


def load_unilist_institutions():
    """Institution names already tracked in ../uni_list.xlsx (read-only)."""
    path = os.path.normpath(os.path.join(
        os.path.dirname(os.path.abspath(__file__)), config.XLSX_PATH))
    names = set()
    try:
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        header = [c.value for c in ws[1]]
        if "University" in header:
            i = header.index("University")
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[i]:
                    names.add(str(row[i]).lower().strip())
    except Exception:
        pass
    return names


def merge(existing, fresh):
    merged = dict(existing)
    for col in HEADERS_XLSX:
        v = fresh.get(col)
        if col in SCRAPER_OWNED and v not in (None, ""):
            merged[col] = v
    merged["New?"] = ""
    for col in HEADERS_XLSX:
        merged.setdefault(col, "")
    return merged


def write_workbook(path, records, new_ids):
    records.sort(key=lambda r: (-(r.get("Score") or 0),
                                str(r.get("Posted") or "")), reverse=False)
    wb = Workbook()
    ws = wb.active
    ws.title = "LinkedIn"
    ws.append(HEADERS_XLSX)
    for c in ws[1]:
        c.fill = HEAD_FILL
        c.font = HEAD_FONT
    for rec in records:
        ws.append([rec.get(h, "") for h in HEADERS_XLSX])
        if str(rec.get("Job ID")) in new_ids:
            for c in ws[ws.max_row]:
                c.fill = NEW_FILL
    for i, h in enumerate(HEADERS_XLSX, 1):
        ws.column_dimensions[get_column_letter(i)].width = \
            COL_WIDTHS.get(h, 16)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS_XLSX))}" \
                         f"{ws.max_row}"
    dv = DataValidation(
        type="list",
        formula1='"Not started,Drafting,Applied,Interview,Offer,Rejected,'
                 'Ignored"', allow_blank=True)
    ws.add_data_validation(dv)
    col = get_column_letter(HEADERS_XLSX.index("Status") + 1)
    dv.add(f"{col}2:{col}{max(ws.max_row, 2)}")
    wb.save(path)


# ------------------------------------------------------------------ main --
def main():
    ap = argparse.ArgumentParser(description="LinkedIn guest-jobs scraper")
    ap.add_argument("--days", type=int, default=7,
                    help="only postings from the last N days (default 7)")
    ap.add_argument("--max-pages", type=int, default=MAX_PAGES_PER_QUERY)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    here = os.path.dirname(os.path.abspath(__file__))
    out_path = os.path.join(here, OUT_XLSX)
    today = dt.date.today().isoformat()

    session = make_session()
    existing = read_existing(out_path)
    uni_names = load_unilist_institutions()
    log(f"{len(existing)} rows already tracked; "
        f"{len(uni_names)} institutions known from uni_list.xlsx")

    # ---- 1. collect job cards across the search grid ----
    found = {}
    for term in SEARCH_TERMS:
        for loc in LOCATIONS:
            for page in range(args.max_pages):
                params = {
                    "keywords": term,
                    "location": loc,
                    "f_TPR": f"r{args.days * 86400}",
                    "start": page * 25,
                }
                r = polite_get(session, SEARCH_URL, params=params)
                if r is None or not r.text.strip():
                    break
                cards = parse_cards(r.text)
                if not cards:
                    break
                for j in cards:
                    found.setdefault(j["id"], j)
                log(f"  '{term}' @ {loc} p{page + 1}: {len(cards)} cards "
                    f"({len(found)} unique so far)")
    log(f"Search grid done: {len(found)} unique postings.")

    # ---- 2. title + country prefilter, then fetch descriptions ----
    candidates = []
    for j in found.values():
        j["country"] = country_of(j["location"])
        if not title_ok(j["title"]):
            continue
        if j["country"] and not country_ok(j["country"]):
            continue
        candidates.append(j)
    log(f"{len(candidates)} pass title+country prefilter.")

    kept, new_ids, fetched = [], set(), 0
    for j in candidates:
        if j["id"] in existing:
            # already tracked: keep, light refresh, no re-fetch
            j["score"] = existing[j["id"]].get("Score") or 0
            j["eligibility"] = existing[j["id"]].get(
                "Eligibility (non-EU)") or ""
            kept.append(j)
            continue
        if fetched >= MAX_DETAIL_FETCHES:
            log("  detail-fetch cap reached; remaining new jobs postponed "
                "to the next run")
            break
        desc = fetch_description(session, j["id"])
        fetched += 1
        j["score"] = score_job(j, desc)
        if j["score"] < MIN_SCORE:
            continue
        j["eligibility"], keep = eligibility_check(desc)
        if not keep:
            log(f"  EXCLUDED ({j['company']}: {j['eligibility']})")
            continue
        new_ids.add(j["id"])
        kept.append(j)
        log(f"  KEEP [{j['score']}] {j['company']} -- {j['title'][:55]}")

    # ---- 3. build records, merge, write ----
    final, seen = [], set()
    for j in kept:
        rec = {
            "Job ID": j["id"], "Source": "LinkedIn",
            "Institution": j["company"], "Position Title": j["title"],
            "Location": j["location"], "Country": j["country"],
            "Posted": j["posted"], "Score": j["score"],
            "Priority": ("High" if j["score"] >= PRIORITY_HIGH else
                         "Medium" if j["score"] >= PRIORITY_MEDIUM
                         else "Low"),
            "Eligibility (non-EU)": j.get("eligibility", ""),
            "Apply Link": j["link"],
            "In uni_list?": ("YES - check for duplicate"
                             if j["company"].lower().strip() in uni_names
                             else ""),
            "Date First Seen": today, "New?": "NEW",
            "Status": "Not started", "Notes": "",
        }
        if j["id"] in existing:
            rec = merge(existing[j["id"]], rec)
        final.append(rec)
        seen.add(j["id"])
    for jid, rec in existing.items():
        if jid not in seen:
            rec = dict(rec)
            for col in HEADERS_XLSX:
                rec.setdefault(col, "")
            final.append(rec)

    log(f"Result: {len(final)} rows, {len(new_ids)} new.")
    if args.dry_run:
        for rec in final:
            if str(rec.get("Job ID")) in new_ids:
                log(f"  NEW [{rec['Score']}] {rec['Institution']} -- "
                    f"{rec['Position Title']} ({rec['Country']})")
        log("Dry run - Excel file not modified.")
        return
    write_workbook(out_path, final, new_ids)
    log(f"Updated {out_path}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        sys.exit("\nInterrupted.")
